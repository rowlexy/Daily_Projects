# Create a multiplication table
# Save the result in multiply.xlsx

import openpyxl

def multiplication_table(output_file):
    try:
        wb = openpyxl.Workbook()
        wb.create_sheet()
        ws = wb.sheetnames # displays sheet names in a list
        print(ws)
        sheet = wb['Sheet1'] # Assigns the sheet name to a variable
        sheet.title = 'Multiplication'
        sheet = wb['Multiplication'] # Re-assigns the sheet name to a variable
        # Checks for duplicate sheet
        for page in ws:
            if len(ws) > 1 and not page == 'Sheet1':
                wb.remove(wb[page])
                
        if len(ws) == 1:
            print('There is only one sheet in the workbook')

        sheet['A1'] = None # Assigns the first cell as a blank value
        
        for row in range(1, 13):
            sheet[f'A{row+1}'].value = row # assigns the value from 'A2'

        for col in range(1, 13):
            sheet.cell(row=1, column=col+1, value=col) #column starts from B1:M1, assigns the value

        for row in range(1, 13):
            for col in range(1, 13):
                result = row * col
                sheet.cell(row=row+1, column=col+1, value=result)
                
        wb.save(output_file) # saves the file
        return True
    
    except PermissionError:
        print('File permission error')
        return False
    except Exception as error:
        print(f'Error, something went wrong {error}')
        return False
    
def main():
    output_file = 'multiplication.xlsx'
    if not output_file.endswith('.xlsx'):
        print('Only excel files with extensions .xlsx are allowed')
    
    read_file = multiplication_table(output_file)
    if read_file:
        print(f'{output_file} successfully created')
    
if __name__ == '__main__':
    main()