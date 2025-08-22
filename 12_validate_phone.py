# Open the vehicles.xlsx Workbook
# Create  a new Header 'Phone number available (Yes / No)
# Change the Station Phone in cell E1 to "Alternate Phone"
# If one or both Phone numbers are available in Station Phone or altenate Phone: 
# Populate the row Phone number available (Yes / No) with "Yes" or 'No as applicable

import openpyxl
from pathlib import Path
# open the workbook
def phone_availabitlity(input_file: str, output_file: str):
    try:
        # if file does not exist
        if not Path(input_file).exists():
            print(f'Error: {input_file} not found')
        # if file is not an excel file
        if not input_file.endswith('.xlsx'):
            print('Only excel files are allowed')
        # load excel file
        wb = openpyxl.load_workbook(input_file)

        # checking if sheetname availability
        if 'vehicles' not in wb.sheetnames:
            print(f"Error: 'vehicles' not found. Available sheet: {wb.sheetnames}")
        ws = wb['vehicles']

        # New header Phone number available (Yes / No)
        ws['H1'] = 'Phone number available (Yes / No)'
        # Change the Station Phone in cell E1 to "Alternate Phone"
        ws['E1'] = 'Alternate Phone'

        for row in range(2, ws.max_row + 1):
            # Cell values of E2 & F2 - end of row
            station_phone = ws[f'D{row}'].value
            alternate_phone = ws[f'E{row}'].value
            
            # if there are no staion phone and alternate phone
            # Populate the cells as 'No
            if station_phone is None and alternate_phone is None:
                ws[f'H{row}'].value = 'No'
            else:
                ws[f'H{row}'].value = 'Yes'
                
        #saving the file        
        wb.save(output_file)
        return True
    
    except FileNotFoundError:
        print(f'Error {input_file} not found')
        return False
    
    except PermissionError:
        print(f'You do not have the permission to view the {output_file} file')
        return False
    
    except Exception as error:
        print(f'Error: something went wrong {error}')
        return False

def main():
    input_file = 'vehicles.xlsx'
    output_file = 'vehicles_edit.xlsx'
    success = phone_availabitlity(input_file, output_file)
    if success:
        print(f'{output_file} created successfully')
        return True
    else:
        print('Process failed, please check errors')
        return False
    
if __name__ == '__main__':
    main()
    