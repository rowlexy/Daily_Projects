# delete the duplicate sheets
# extract the data from sheet1
# save the information as a json file
# Transfer the data 
# Create a new excel file
# Add % difference, Average & Difference
# 
import openpyxl, json, os
from pathlib import Path
from taxes_dict import taxes_data
def removeDuplicateSheets(input_file, output_file):
    try:
        if not Path(input_file).exists():
            raise FileNotFoundError(f'Error: {input_file} not found')
        wb = openpyxl.load_workbook(input_file)
        sheetname = wb.sheetnames
        print(sheetname)

        if len(sheetname) > 0:
            sheet_to_delete = sheetname[1:]
            for sheet in sheet_to_delete:
                wb.remove(wb[sheet])
                print(f'{sheet} deleted')
        wb.save(output_file)
        return True
    except Exception as e:
        print(f'Error: {e}')
        return False
if removeDuplicateSheets('taxes.xlsx', 'taxes_edit.xlsx'):
    print('File created successfully')
    
def convertExcelToJSON(input_file, output_file):
    try:
        if not os.path.exists(input_file):
            raise FileNotFoundError(f'Error: {input_file} not found')
        wb = openpyxl.load_workbook(input_file)
        ws = wb['Sheet1']
        order = ws['A1'].value
        country = ws['B1'].value
        gas = ws['C1'].value
        diesel = ws['D1'].value
        headers = [order, country, gas, diesel]
        # create a list to hold our JSON data
        data = []
        for row in range(2, ws.max_row + 1):
            data_to_json = {
                headers[0]: ws[f'A{row}'].value,
                headers[1]: ws[f'B{row}'].value,
                headers[2]: ws[f'C{row}'].value,
                headers[3]: ws[f'D{row}'].value
            }
            data.append(data_to_json)
        with open(output_file, 'w') as json_data:
            json.dump(data, json_data, indent=2)
        return data
    except Exception as e:
        print(f'Error: {e}')
        return False
if convertExcelToJSON('taxes_edit.xlsx', 'data.json'):
    print('file successfully converted to JSON')
    
def newExcelFile(output_file):
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='Taxes')
    sheet = wb['Taxes']
    column_header = ['Order', 'Country', 'Gasoline (road use) $/gal', 'Diesel (road use) $/gal']
    sheet.append(column_header)
    for info in taxes_data:
        order = info['Order']
        country = info['Country']
        gasoline = info['Gasoline (road use) $/gal']
        diesel = info['Diesel (road use) $/gal']
        result = [order, country, gasoline, diesel]
        sheet.append(result)
        
    sheet['E1'] = 'Average Gasoline & Diesel'
    sheet['F1'] = 'Difference Gasoline & Diesel'
    sheet['G1'] = 'Percentage Difference Gasoline & Diesel'
    
    for row in range(2, sheet.max_row + 1):
        gasoline_price = round(sheet[f'C{row}'].value, 3)
        diesel_price = round(sheet[f'D{row}'].value, 3)
        difference = round(gasoline_price - diesel_price, 3)
        sheet[f'F{row}'] = difference
        average = round((gasoline_price + diesel_price) / 2, 2)
        if average != 0:
            sheet[f'E{row}'] = average
            percentage_diff = round((difference / average) * 100, 2)
            sheet[f'G{row}'] = percentage_diff
        else:
           sheet[f'E{row}'] = 0
           sheet[f'G{row}'] = 0
        
    wb.save(output_file)
    
if newExcelFile('taxes.xlsx'):
    print('File created successfully')